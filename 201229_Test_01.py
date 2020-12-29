import openpyxl

from openpyxl import Workbook

wb =  openpyxl.load_workbook ('201227_DataSpread-JY.xlsx', data_only = True)

#user greeting
name = input("What's your name? ")
print("It's nice to meet you,", name)

print(name,',','The data base you will be using today is', wb)

#checking sheet names
# sn = wb.sheetnames
# print(sn)

print("Does your development exist within a void? If so, which?")
options = ["Void of Digital Connection", "Void of Urban Connectivity", "Void of Flood", "voids of urban Vulnerability", "Void of Shock Absorbtion","No Void"]



boo2 = ""
boo3 = ""
boo4 = ""
inp3 = ""
inp4 = ""
inp5 = ""

def checkboo2():
    global boo2, inp3
    boo2 = input("Is there another void on site? Y for yes / N for no")
    if boo2 == "Y":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        inp3 = int(input("Enter a number: "))
        if inp3 == 1:
            print("you have picked", inp3, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Technology + information' category.")
        elif inp3 == 2:
            print("you have picked", inp3, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
        elif inp3 == 3:
            print("you have picked", inp3, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Water', 'Green Infratsture' and 'Shelter' Categories.")
        elif inp3 == 4:
            print("you have picked", inp3, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Building and Design','Economy'and 'Social and Demographic' categories.")
        elif inp3 == 5:
            print("you have picked", inp3, ", on top of your previous void(s), your development is expected to have a higher impact in the 'water','planning', 'Connectivity'and 'Social and Demographic' categories.")
        else:
            print("Invalid input!")
    elif boo2 == "N":
        for i in range(len(options)):
                print(str(i+1) + ":", options[i])
        print("Your Final selection is", inp1, "and", inp2)
    else:
        print("Invalid input!")

def checkboo3():
    global boo3, inp3, inp4
    boo3 = input("Is there another void on site? Y for yes / N for no")
    if boo3 == "Y":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        inp4 = int(input("Enter a number: "))
        if inp4 == 1:
            print("you have picked", inp4, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Technology + information' category.")
        elif inp4 == 2:
            print("you have picked", inp4, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
        elif inp4 == 3:
            print("you have picked", inp4, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Water', 'Green Infratsture' and 'Shelter' Categories.")
        elif inp4 == 4:
            print("you have picked", inp4, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Building and Design','Economy'and 'Social and Demographic' categories.")
        elif inp4 == 5:
            print("you have picked", inp4, ", on top of your previous void(s), your development is expected to have a higher impact in the 'water','planning', 'Connectivity'and 'Social and Demographic' categories.")
        else:
            print("Invalid input!")
    elif boo3 == "N":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        print("Your Final selection is", inp1, "and", inp2, "and", inp3)
    else:
        print("Invalid input!")

def checkboo4():
    global boo4, inp3, inp4, inp5
    boo4 = input("Is there another void on site? Y for yes / N for no")
    if boo4 == "Y":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        inp5 = int(input("Enter a number: "))
        if inp5 == 1:
            print("you have picked", inp5, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Technology + information' category.")
        elif inp5 == 2:
            print("you have picked", inp5, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
        elif inp5 == 3:
            print("you have picked", inp5, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Water', 'Green Infratsture' and 'Shelter' Categories.")
        elif inp5 == 4:
            print("you have picked", inp5, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Building and Design','Economy'and 'Social and Demographic' categories.")
        elif inp5 == 5:
            print("you have picked", inp5, ", on top of your previous void(s), your development is expected to have a higher impact in the 'water','planning', 'Connectivity'and 'Social and Demographic' categories.")
        else:
            print("Invalid input!")
        print("Your Final selection is", inp1, "and", inp2, "and", inp3,"and", inp4, "and", inp5)
    elif boo4 == "N":
        for i in range(len(options)):
            print(str(i+1) + ":", options[i])
        print("Your Final selection is", inp1, "and", inp2, "and", inp3,"and", inp4)
    else:
        print("Invalid input!")

# Print out options
for i in range(len(options)):
    print(str(i+1) + ":", options[i])

# Initial question on void condition within development
inp1 = int(input("Enter a number: "))
if inp1 == 1:
    print("you have picked", inp1, ", your development is expected to have a higher impact in the 'Technology + information' category.")
elif inp1 == 2:
    print("you have picked", inp1, ", your development is expected to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
elif inp1 == 3:
    print("you have picked", inp1, ", your development is expected to have a higher impact in the 'Water', 'Green Infratsture' and 'Shelter' Categories.")
elif inp1 == 4:
    print("you have picked", inp1, ", your development is expected to have a higher impact in the 'Building and Design','Economy'and 'Social and Demographic' categories.")
elif inp1 == 5:
    print("you have picked", inp1, ", your development is expected to have a higher impact in the 'water','planning', 'Connectivity'and 'Social and Demographic' categories.")
elif inp1 == 6:
    print("you have picked", inp1, ", your development is not expected to have a higher impact in a specific area .")
else:
    print("Invalid input!")


# Asking user if there is another void on site
boo1 = input("Is there another void on site? Y for yes / N for no")
if boo1 == "Y":
    for i in range(len(options)):
        print(str(i+1) + ":", options[i])
    inp2 = int(input("Enter a number: "))
    if inp2 == 1:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Technology + information' category.")
    elif inp2 == 2:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Connectivity','Technology + Information'and 'Spatial Configuration' categories.")
    elif inp2 == 3:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Water', 'Green Infratsture' and 'Shelter' Categories.")
    elif inp2 == 4:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'Building and Design','Economy'and 'Social and Demographic' categories.")
    elif inp2 == 5:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'water','planning', 'Connectivity'and 'Social and Demographic' categories.")
    elif inp2 == inp1:
        print("you have picked", inp2, ", on top of your previous void(s), your development is expected to have a higher impact in the 'water','planning', 'Connectivity'and 'Social and Demographic' categories.")
    else:
        print("Invalid input!")
elif boo1 == "N":
    for i in range(len(options)):
        print(str(i+1) + ":", options[i])
    print("Your Final selection is", inp1)
else:
    print("Invalid input!")


if boo1 != "N":
    checkboo2()
    if boo2 != "N":
        checkboo3()
        if boo3 != "N":
            checkboo4()

            
# Asking user if there is another void on site

    
# Asking user if there is another void on site


# sheet = wb['Sheet1']

# sheet['D5'].value 

# print(sheet['D5'].value)

#for i in range(4,83):print(sheet.cell(row=i,column=4).value)
# print(sheet.max_row)

# wb.create_sheet(title='My Sheet Name', index=1)

# wb.save ('example.xlsx')
